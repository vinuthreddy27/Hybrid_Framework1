import openpyxl

def get_list_from_excel(excel_name,sheet_name):
    excel=openpyxl.load_workbook(excel_name)
    sheet=excel[sheet_name]
    values=[]
    for row in range(1,sheet.max_column+1):
        nested_creds=[]
        for col in range(1,sheet.max_column+1):
            data=sheet.cell(row,col)
            nested_creds.append(data.value)
        values.append(nested_creds)

    return values


